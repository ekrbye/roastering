# -*- coding: utf-8 -*-
"""
Created on Sat Aug  5 18:14:09 2023

@author: wongy
"""

import pandas as pd
import numpy as np
from default import *
import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class Transposer():
    def __init__(self):
        self.df_work = None

    def num_tofield(self, num: (int, list), field) -> np.ndarray:
        if isinstance(num, int): num = [num]
        return self.df_work.iloc[num,][field].to_numpy()

    def extract_names(text: str) -> list:
        """
        将text的中文名字提取，其中text由空格、'[]'分隔

        Parameters
        ----------
        text : TYPE
            DESCRIPTION.

        Returns
        -------
        names : TYPE
            DESCRIPTION.

        """
        # 使用正则表达式匹配中文字符
        pattern = r'[\u4e00-\u9fa5]+'
        # 使用re.split()函数根据空格或[]分割字符串
        text = re.split(r'\s|\[|\]', text)
        # 提取所有匹配的中文名字
        names = [re.findall(pattern, name) for name in text if re.findall(pattern, name)]
        # 将嵌套的列表展平
        names = [name for sublist in names for name in sublist]
        return names
    
    def delete_num(self, num: int, dict_dep: dict, pdomain: (list, tuple), 
                   copy=False) -> dict:
        """
        删除指定pdomain的num

        Parameters
        ----------
        num : int
            DESCRIPTION.
        dict_dep : dict
            DESCRIPTION.
        pdomain : (list, tuple)
            DESCRIPTION.
        copy : TYPE, optional
            DESCRIPTION. The default is False.

        Returns
        -------
        dict
            DESCRIPTION.

        """
        if copy:
            _dict_dep = dict_dep.copy()
        else:
            _dict_dep = dict_dep
        for p, l_v in dict_dep.items():
            if p[1] in pdomain:
                dict_dep[p] = list(filter(lambda x: x != num, l_v))
    
    def dep_tosheet(self, dict_dep: dict, field: str=STR_SHEETFIELD_NAME) -> pd.DataFrame:
        """
        将dict_dep信息转换成可读的内容，返回DataFrame

        Parameters
        ----------
        dict_dep : dict
            DESCRIPTION.
        field : str, optional
            DESCRIPTION. The default is STR_SHEETFIELD_NAME.

        Returns
        -------
        df : TYPE
            DESCRIPTION.

        """
        len_rows = max(dict_dep.keys())[0] + 1
        len_columns = max(dict_dep.keys())[1] + 1
        l_rows = LIST_SHEETCOLUMNS
        l_columns = DICT_INDEXDEP.values()
        df = pd.DataFrame(np.empty((len_rows, len_columns)), 
                          columns=l_rows, 
                          index=l_columns)
        fr = lambda s: re.sub(r"[\[\]']", "", s)
        for r in range(len_rows):
            for c in range(len_columns):
                text = str(self.num_tofield(dict_dep[(r, c)], field))
                df.iloc[r, c] = fr(text)
        return df
    
    
    def save_dftoxlsx(self, df: pd.DataFrame, filename: str='output') -> None:
        """
        保存df为xlsx，文件名为filename，df由dep_tosheet生成

        Parameters
        ----------
        df : pd.DataFrame
            DESCRIPTION.
        filename : str, optional
            DESCRIPTION. The default is 'output'.

        Returns
        -------
        None
            DESCRIPTION.

        """
        
        with pd.ExcelWriter(filename + '.xlsx', engine='xlsxwriter', 
                            options={'strings_to_urls': False}) as writer:
            df.to_excel(writer, index=True, encoding='utf-8')

class Sheet():
    def __init__(self, filename):
        self.workbook = openpyxl.load_workbook(filename + ".xlsx")
        self.worksheet = self.workbook.active
    
    def set_colwidth(self, 
                     w: int =INT_SHEETCOLUMNWIDTH, 
                     col: (int, list, tuple) =None) -> None:
        if col is None:
            col = [get_column_letter(i + 1) for i in range(len(LIST_SHEETCOLUMNS) + 1)]
        if isinstance(col, int): col = [col]
        for i in col:
            self.worksheet.column_dimensions[i].width = w
    
    def set_rowheight(self, 
                      h: int =INT_SHEETROWHEIGHT, 
                      row: (int, list, tuple) =None) -> None:
        if row is None:
            row = [i + 1 for i in range(len(DICT_INDEXDEP.keys()) + 1)]
        if isinstance(row, int): row = [row]
        for i in row:
            self.worksheet.row_dimensions[i].height = h
    """
    def set_font(self, 
                 font_style: dict, 
                 row: (int, list, tuple) =None, 
                 col: (int, list, tuple) =None) -> None:
        if isinstance(row, int): row = [row]
        if isinstance(col, int): col = [col]
        for r in row:
            for c in col:
                cell = self.worksheet.cell(row=r, column=c)
                cell.font = Font(**font_style)
    """
    def set_font(self, 
                 font_style: dict, 
                 row: (int, list, tuple) =None, 
                 col: (int, list, tuple) =None) -> None:
        self.set_cellalignment(row=row, col=col, font=Font(**font_style))
    
    def set_cellalignment(self,
                 row: (int, list, tuple) =None, 
                 col: (int, list, tuple) =None, 
                 **kwargs) -> None:
        """
        设置行数为row，列数为col的单元格属性，注意表格行列从1开始

        Parameters
        ----------
        row : (int, list, tuple), optional
            DESCRIPTION. The default is None.
        col : (int, list, tuple), optional
            DESCRIPTION. The default is None.
        **kwargs : TYPE
            DESCRIPTION.

        Returns
        -------
        None
            DESCRIPTION.

        """
        if isinstance(row, int): row = [row]
        if isinstance(col, int): col = [col]
        for r in row:
            for c in col:
                cell = self.worksheet.cell(row=r, column=c)
                self.set_alignment(cell, **kwargs)
    
    def set_alignment(self, cell, **kwargs):
        for k, v in kwargs.items():
            setattr(cell, k, v)
                
    def set_wrapalignment(self, 
                          wrapalignment_style: dict, 
                          row: (int, list, tuple) =None, 
                          col: (int, list, tuple) =None) -> None:
        self.set_cellalignment(row=row, col=col, alignment=Alignment(**wrapalignment_style))
        
    def _set_default(self):
        col = [i + 1 for i in range(len(LIST_SHEETCOLUMNS) + 1)]
        row = [i + 1 for i in range(len(DICT_INDEXDEP.keys()) + 1)]
        self.set_font(row=1, col=col, 
                      font_style=DICT_FONTPARAMETERS_TITLETEXT)
        self.set_font(row=row, col=1, 
                      font_style=DICT_FONTPARAMETERS_TITLETEXT)
        self.set_font(row=row[1: ], col=col[1: ], 
                      font_style=DICT_FONTPARAMETERS_MAINTEXT)
        self.set_wrapalignment(row=row, col=col, 
                               wrapalignment_style=DICT_ALIGNMENTPARAMETERS_MAINTEXT)
        """
        wrap_alignment = Alignment(horizontal='center', 
                                     vertical='center', 
                                     wrap_text=True)
        for r in row:
            for c in col:
                cell = self.worksheet.cell(row=r, column=c)
                cell = self.worksheet.cell(row=r, column=c)
                cell.alignment = wrap_alignment
        """